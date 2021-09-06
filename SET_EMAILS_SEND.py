import smtplib
import time
import openpyxl
import os

###################################################################

subjectOfEmail = ""
while not subjectOfEmail:
    subjectOfEmail = input("Enter the Subject of the email: ")

print("\n\n")

###################################################################

os.chdir(r'C:\\Users\\abdul\\my-python-scripts')
wb = openpyxl.load_workbook('names and emails.xlsx')
sheet = wb['Sheet1']
receivers = {}
row = 2
while True:
    name = sheet.cell(row=row, column=1).value
    email = sheet.cell(row=row, column=2).value

    if email:
        receivers[email] = name
    else:
        break
    row = row + 1

if not receivers:
    print("No Data found in the names and email excel file.")
    quit()

print("Enter your email in the notepad file")
time.sleep(1)

print("\n\n")

###################################################################

emailMessageFileName = 'email_message.txt'
os.chdir(r'C:\\Users\\abdul\\my-python-scripts')
os.system(f'notepad {emailMessageFileName}')

file = open(emailMessageFileName)
message = file.read()
file.close()

###################################################################

port = 465
connection = smtplib.SMTP('smtp.office365.com', 587)  # For gmail replace office365 with gmail
connection.ehlo()
connection.starttls()

condition = True
while condition:
    try:
        senderEmail = 'l1f18bscs0324@ucp.edu.pk'  # Replace this email with your email
        senderPassword = ""

        while not senderPassword:
            senderPassword = input(f"Enter Password for email {senderEmail}: ")

        connection.login(senderEmail, senderPassword)
        condition = False
        print("\n\n")
    except:
        print("Wrong Password\n")

for email in receivers:
    output = connection.sendmail(senderEmail, email,
                                 f'Subject:{subjectOfEmail}\n\nDear {receivers[email] if receivers[email] else ""},\n\n{message}')

    if output != {}:
        print(f"Email not sent to {email}")
    else:
        print(f"Email sent successfully to {email}")

connection.quit()
